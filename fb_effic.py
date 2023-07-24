import pandas as pd
import time
import os
import numpy as np
from openpyxl import Workbook, load_workbook
import streamlit as st
from io import BytesIO
import streamlit_ext as ste
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.utils.cell

st.set_page_config(page_title='First Bus Optibus Scripts')
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

class AggrCrewSchedule():
    def __init__(self, file_path = None):
        #capture file path
        self.path = file_path

        self.result = None

        #Get the sheet names from the file
        wb = load_workbook(self.path)

        df = pd.read_excel(self.path, sheet_name = wb.sheetnames[0])

        _columnNamesList = list(df)
        _columnNamesList.append('Service Id')

        _serviceIdCorrespDict = {}

        for _name in wb.sheetnames:
            _correspName = _name
            if "mon" in _name.lower() or "fri" in _name.lower() or ("sat" not in _name.lower() and "sun" not in _name.lower()):
                _correspName = 'M-F'
            _serviceIdCorrespDict[_name] = _correspName

        _concatDataFrame = pd.DataFrame(columns = _columnNamesList)
        #At this point we have an empty final dataFrame that will be used to create the dictionaries for the duty types

        for sheet in wb.sheetnames:
            
            df = pd.read_excel(self.path, sheet_name = sheet, dtype=str)
            df['Service Id'] = _serviceIdCorrespDict[sheet]
            _concatDataFrame = pd.concat([_concatDataFrame, df])
        
        self.dataFrame =  _concatDataFrame
        #Now we have the merged dataframes


class TimeString_HHMM():
    def __init__(self, input_string = None):
        #split at position 0 up to first colon and fill 2 0s - HOURS 
        self.hh = input_string.split(':')[0].zfill(2)
        #split at position 1 up to first colon and fill 2 0s - MINUTES
        self.mm = input_string.split(':')[1].zfill(2)
        #Convert hours to minutes to get total minutes value by hour-minutes + minutes
        self.in_minutes = int(self.hh)*60 + int(self.mm)

class FullSchedule():
    def __init__(self, file_path = None, aggrFilePath = None):
        #Aggregated Crew Schecule
        self.crewSchedule = AggrCrewSchedule(aggrFilePath)
        #The AggrCrewSchedule DataFrame is accessible at self.crewSchedule.dataFrame
        
        #filepath
        self.path = file_path

        #Create a DataFrame out of the raw Full Schedule
        df = pd.read_excel(self.path, dtype=str)

        columnsToCheck = ['Duty id', 'Event Type', 'Route Id', 'Sign', 'Direction',
                  'Alternative', 'Start Time', 'End Time', 'Origin Stop Id',
                  'Destination Stop Id', 'Days']

        df.drop_duplicates(subset=columnsToCheck, keep='first', inplace=True)

        df['Service Group Days'] = df['Days']

        df = df.reset_index(drop=True)

        #replace blanks with nans in event type column
        df['Event Type'].replace('', np.nan, inplace=True)
        #drop na's in event type
        df.dropna(subset=['Event Type'], inplace=True)

        self.dataFrame = df
        #create a list of service group days by the service groupe days column in the list (UNIQUE BASED ON SET)
        self.serviceGroupDaysList = list(set(self.dataFrame['Service Group Days'].to_list()))
        #create event type list by getting event type (UNIQUE BASED ON SET)
        self.eventTypeList = list(set(self.dataFrame['Event Type'].to_list()))
        #create a list from preference groups (UNIQUE BASED ON SET)
        self.prefGroups = list(set(self.dataFrame['Pref Group'].to_list()))


    #Adapt the DataFrame the way the client did in Excel 
    def insert_extra_columns(self):

       

        df = self.dataFrame

        # Initialize st.session_state['rule_list'] if it does not exist
        if 'rule_list' not in st.session_state:
            st.session_state['rule_list'] = []

        # Standard rules
        standard_rule = st.expander('**Standard Rules**', expanded=True)
        mins_unpaid = standard_rule.number_input('Standard unpaid break time (mins)', min_value=0, max_value=90, value=60, step=5)
        mins_unpaid_dec = mins_unpaid / 60
        standard_makeup_val = standard_rule.number_input('Standard Guarantee Time (hrs)', min_value=4, max_value=9, value = 6)

        # Custom rule form
        custom_rule = st.expander('*Custom rules*', expanded=True)

        # Create a selectbox to select the rule type
        rule_type = custom_rule.selectbox('Select a rule type', ['Break Exception', 'Makeup Exception'])

        if rule_type == 'Break Exception':
            with custom_rule.form("Break Rules"):
                rule_group = st.selectbox("Select a pref group", df['Pref Group'].unique())
                rule_days = st.multiselect("Select days of the week", df['Service Group Days'].unique())
                rule_break_duration = st.number_input('Enter the unpaid break duration in minutes', min_value=0, max_value=90, step=5, value=30)/60
                rule_add_button_clicked = st.form_submit_button("Add Break Rule")

                if rule_add_button_clicked:
                    new_rule = {'type': 'break', 'group': rule_group, 'days': rule_days, 'break_duration': rule_break_duration}
                    st.session_state['rule_list'].append(new_rule)

        elif rule_type == 'Makeup Exception':
            with custom_rule.form("Makeup Rules"):
                rule_group = st.selectbox("Select a pref group", df['Pref Group'].unique())
                rule_days = st.multiselect("Select days of the week", df['Service Group Days'].unique())
                rule_makeup_val = st.number_input('Enter the Guarantee Time (hrs)', min_value=0, max_value=9, value=6)
                rule_add_button_clicked = st.form_submit_button("Add Makeup Rule")

                if rule_add_button_clicked:
                    new_rule = {'type': 'makeup', 'group': rule_group, 'days': rule_days, 'makeup_val': rule_makeup_val}
                    st.session_state['rule_list'].append(new_rule)

        

        
        for i, rule in enumerate(st.session_state['rule_list']):
            if rule['type'] == 'break':
                rule_info = f"Rule {i+1}: {rule['type']}, Pref group = {rule['group']}, Days of week: **{', '.join(rule['days'])}**, Unpaid break duration: **{rule['break_duration']*60} mins**"
            elif rule['type'] == 'makeup':
                rule_info = f"Rule {i+1}: {rule['type']}, Pref group = {rule['group']}, Days of week: **{', '.join(rule['days'])}**, Guarantee Time = **{rule['makeup_val']} hrs**"
            
            col1, col2 = st.columns([4,1]) # adjust the values for your layout
            with col1:
                st.info(rule_info)
            with col2:
                if st.button(f"Delete Rule {i+1}"):
                    del st.session_state['rule_list'][i]
       
# Filter the dataframe based on the selected groups

    
        #create new columns for the dataframe using empty lists
        _TimeColumn = []
        _MeasureColumn = []
        _LayoverJoinUpColumn = []
        _PaidColumn = []
        _BreakCountColumn = []

        #Mapping for Break Count
        #creating a mapping dataframe where copying standby event type only
        mappingDf = df[df['Event Type'] == 'standby'].copy()
        #Creating a new column, concatentating duty ID and mapping df with underscore
        mappingDf['mappingColumn'] = mappingDf['Duty id'] + '_' + mappingDf['Service Group Days']

        #Use this logic for creating joinup time rule 
        df['mappingColumn'] = df['Duty id'] + '_' + df['Service Group Days']
        ##################################
        _mappingList = mappingDf['mappingColumn'].to_list()

        for index, row in df.iterrows():
            startTime = TimeString_HHMM(row['Start Time'])
            endTime = TimeString_HHMM(row['End Time'])
            endHour = endTime
            
            if int(startTime.hh) > int(endTime.hh):
                endHour = TimeString_HHMM(f"{int(endTime.hh) + 24}:{endTime.mm}")
            
            duration = (endHour.in_minutes - startTime.in_minutes)*24*0.000694 #This is the factor that Excel uses for the Time column
            _TimeColumn.append(duration)

            if row['Event Type'] == 'standby' and (row['Description'] == 'Break' or row['Description'] == 'Paid Break'):

                

                _matchingCode = f"{row['Duty id']}_{row['Service Group Days']}"
                _breakCount = _mappingList.count(_matchingCode)

               

                custom_unpaid_value = None
                for rule in st.session_state['rule_list']:
                    if rule['type']=='break':
                        if row['Pref Group'] == rule['group'] and row['Service Group Days'] in rule['days']:
                            custom_unpaid_value = rule['break_duration']
                            break

                if custom_unpaid_value is None:
                    _timePaidTime = duration
                    if duration < mins_unpaid_dec:
                        _paidTimeValue = 0.00
                    else:
                        _paidTimeValue = duration - mins_unpaid_dec
                else:
                    _timePaidTime = duration
                    if duration < custom_unpaid_value:
                        _paidTimeValue = 0.00
                    else:
                        _paidTimeValue = duration - custom_unpaid_value


            else:
                #all other events whucg don't meet the row[eventtype] condition are unpaid
                _paidTimeValue = 0.00
                _breakCount = 0
            
            _PaidColumn.append(_paidTimeValue)
            _BreakCountColumn.append(_breakCount)


            if index == 0:
                _MeasureColumn.append('JOINUP')
                _LayoverJoinUpColumn.append(0.00)

            else:
                currentRow_eventType = row['Event Type']
                previousRow_eventType = df.iloc[index-1]['Event Type']

                if currentRow_eventType == previousRow_eventType or (currentRow_eventType == 'service_trip' and previousRow_eventType == 'depot_pull_out') or (currentRow_eventType == 'service_trip' and previousRow_eventType == 'deadhead'):
                    _MeasureColumn.append('Layover')
                else:
                    _MeasureColumn.append('JOINUP')
                
                currentRow_dutyId = row['Duty id']
                previousRow_dutyId = df.iloc[index-1]['Duty id']

                currentRow_dutyIdx = row['mappingColumn']
                previousRow_dutyIdx = df.iloc[index-1]['mappingColumn']

                currentRow_startTime = TimeString_HHMM(row['Start Time'])
                previousRow_endTime = TimeString_HHMM(df.iloc[index-1]['End Time'])
                _startHour = currentRow_startTime

                if int(_startHour.hh) < int(previousRow_endTime.hh):
                    #Surely this should now be adding the value to previous row end time but current row start hour? 
                    #_startHour = TimeString_HHMM(f"{int(previousRow_endTime.hh) + 24}:{previousRow_endTime.mm}")

                    #THIS HAS GOT THE MAX VALUE DISTRIBUTION DOWN FOR THE LAYOVER VALUES AFTER MIDNIGHT
                    _startHour = TimeString_HHMM(f"{int(currentRow_startTime.hh) + 24}:{currentRow_startTime.mm}")
                #Used this to ensure matching service days 
                #if currentRow_dutyId == previousRow_dutyId:
                if currentRow_dutyIdx == previousRow_dutyIdx:
                    _value = (_startHour.in_minutes - previousRow_endTime.in_minutes)*24*0.000694
                    if _value < 0:
                        _value = 0.00
                else:
                    _value = 0.00
                
                _LayoverJoinUpColumn.append(_value)

        #TODO: IMPLEMENT CUSTOM BREAK RULE FOR CUMULATIVE BREAK 

       

                #_LayoverJoinUpColumn = [round(item, 2) for item in _LayoverJoinUpColumn]

        df['Time'] = _TimeColumn
        df['Measure'] = _MeasureColumn
        df['Layover / Join Up Value'] = _LayoverJoinUpColumn
        df['Paid'] = _PaidColumn
        df['Break Count'] = _BreakCountColumn

        df['matchingKeyBreakCount'] = df['Duty id'] + '_' + df['Service Group Days']

        _auxDf = df.copy()
        _auxDf = _auxDf[_auxDf['Event Type'] == 'standby'].copy()

        _auxDf = _auxDf.sort_values('Time', ascending=False).drop_duplicates('matchingKeyBreakCount').sort_index()

        _auxDf.set_index('matchingKeyBreakCount', inplace=True)
        _BreaksCountDict = _auxDf.to_dict('index')

        _newPaidColumn = [] #Storing the new Paid values in here following the client's rules

        for index, row in df.iterrows():
            if int(row['Break Count']) <= 1:
                _newPaidColumn.append(row['Paid'])
            else:
                if row['Start Time'] == _BreaksCountDict[row['matchingKeyBreakCount']]['Start Time']:
                    _newPaidColumn.append(row['Paid'])
                else:
                    _newPaidColumn.append(row['Time'])

        df['newPaid'] = _newPaidColumn

        self.adaptadedDataFrame = df

        overlap_list = list(df[df['Layover / Join Up Value'] > 2].index.values)

        df['Layover / Join Up Value'][df['Layover / Join Up Value']>2] = 0


        df['new_time'] = df.apply(lambda row: row['Paid'] if row['Event Type'] in ['standby', 'split'] else row['Time'], axis=1)

        # Create the paid time sum which we can use to get the joinup values for the summary data
        df['time_sum'] = df['new_time'] + df['Paid'] + df['Layover / Join Up Value'] + df['newPaid']


        df['makeup'] = 0
        # loop through each unique matchingKeyBreakCount group
        for group_name, group_df in df.groupby('matchingKeyBreakCount'):
            # Find a matching custom rule if it exists
            makeup_val = standard_makeup_val
            group, day = group_name.split('_', 1)
            for rule in st.session_state['rule_list']:
                if rule['type'] == 'makeup' and group == rule['group'] and day in rule['days']:
                    makeup_val = rule['makeup_val']
                    break

            # calculate the sum of time_sum for this group
            sum_time_sum = group_df['time_sum'].sum()
            # if sum_time_sum is less than makeup_val, calculate makeup value and apply to last row in group
            if sum_time_sum < makeup_val:
                makeup = makeup_val - sum_time_sum
                last_row_index = group_df.index[-1]
                df.at[last_row_index, 'makeup'] = makeup

            

            #create a new column for makeup, 
            # the sum of time_sum for 'matchingKeyBreakCount' if this is less than 6, then 6 - sum(time_sum) for matchingKeyBreakCount, put this value to the first occurunce of the row and then rest is 0 until next matchingKeyBreakCount, 

            # All values in defined prefGroup 0 . 

       

    # guarantee time in mins
        

        overlap_pairs_list = []

        for index in overlap_list:
            overlap_pairs_list.append(int(index)-1)
            overlap_pairs_list.append(int(index))



        pairs_df = df.copy()
        pairs_df = pairs_df[pairs_df.index.isin(overlap_pairs_list)]

  
        #dfpaid = df[df['Event Type'] == 'standby'].copy()
        
    
        buffer = BytesIO()
        with pd.ExcelWriter(buffer) as writer:
            df.to_excel(writer, sheet_name='Crew Schedule')
            pairs_df.to_excel(writer, sheet_name='Overlapping Trip Pairs')
        ste.download_button('Download Calculated Dataframe', data = buffer, file_name=f'{file_name}_calc_crew_schedule.xlsx')
        


        
        

        
    ########################################################################################################################################

        #CREATE A FILTERED DATAFRAME TO LOOK FOR DISCEPENCIES

  


            #MAX VALUE OF JOINUP BEING SCEWED DUE TO DEADHEAD OVERLAPPING WITH TRIP TIME: 
            #07:40 - 08:04 (start time, end time) - 5452 idx - Deadhead 
            #07:40 - 08:49 (start time, end time) - 5452 idx - Trip

            #7012:7013

        #st.write(df['Layover / Join Up Value'].sum())


    def build_printable_table(self, pref_group = None):
        if pref_group == None:
            pref_group = self.prefGroups
        elif type(pref_group) == str:
            pref_group = [pref_group]
        
        #THESE STRINGS WEREN"T SEPERATED SO SEPERATED TO CALCULATE ACCORDINGLY
        _signList = ['RF', 'RH', 'SHTL']
        
        df = self.adaptadedDataFrame
        df = df[df['Pref Group'].isin(pref_group)] #Here we already filter for the desired preference groups
        timePerEvetTypeDict = get_time_per_event_type_dict(df, self.eventTypeList, self.serviceGroupDaysList)
        
        timePerEvetTypeDict = add_missing_event_types(timePerEvetTypeDict)
        
        timePerEvetTypeDict = check_subkeys(timePerEvetTypeDict)
        
       
        
        

        AggrCrewScheduleDataFrame = self.crewSchedule.dataFrame
    
        paidTimeDict = get_pivot_sheet2(df, self.serviceGroupDaysList)
        layoverJoinUpDict = get_pivot_sheet5(df, self.serviceGroupDaysList)
        layoverJoinUpDict = check_subkeys(layoverJoinUpDict)
        reliefTimeDict = get_sum_of_time_relief_cars(df, self.serviceGroupDaysList, _signList)
        reliefTimeDict = check_keys(reliefTimeDict)
        dutyTypeDict = get_duty_type_number_dict(AggrCrewScheduleDataFrame)
        dutyTypeDict = check_subkeys_weekdays(dutyTypeDict)

        makeupTimeDict = get_makeup_per_duty_event(df, self.serviceGroupDaysList)

       
        paidTimeDict = check_keys(paidTimeDict)
        
        #Build a new Dict for the final table plotting
        finalTableDict = {}
        finalTableDict['Pref Group'] = pref_group

        #In service Time
        #download calculated dataframe
        
        # st.write(paidTimeDict)

        

       

        #WHY IS relieftime dict only getting 0's - need to check this function is correct
      

        
        finalTableDict['In service Time'] = {}
        finalTableDict['In service Time']['M-F'] = _monFri = round(timePerEvetTypeDict['service_trip']['23456'] - reliefTimeDict['23456'], 2)
        finalTableDict['In service Time']['Sat'] = _sat = round(timePerEvetTypeDict['service_trip']['7'] - reliefTimeDict['7'],2)
        finalTableDict['In service Time']['Sun'] = _sun = round(timePerEvetTypeDict['service_trip']['1'] - reliefTimeDict['1'],2)
        finalTableDict['In service Time']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['In service Time']['Comments'] = 'In service driving time'

        
        
        

        #Deadtime
        finalTableDict['Deadtime'] = {}
        finalTableDict['Deadtime']['M-F'] = _monFri = round(timePerEvetTypeDict['deadhead']['23456'] + timePerEvetTypeDict['depot_pull_in']['23456'] + timePerEvetTypeDict['depot_pull_out']['23456'],2)
        finalTableDict['Deadtime']['Sat'] = _sat = round(timePerEvetTypeDict['deadhead']['7'] + timePerEvetTypeDict['depot_pull_in']['7'] + timePerEvetTypeDict['depot_pull_out']['7'],2)
        finalTableDict['Deadtime']['Sun'] = _sun = round(timePerEvetTypeDict['deadhead']['1'] + timePerEvetTypeDict['depot_pull_in']['1'] + timePerEvetTypeDict['depot_pull_out']['1'],2)
        finalTableDict['Deadtime']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Deadtime']['Comments'] = 'Positioning trips time'

        #Attendance recovery time
        finalTableDict['Attendance'] = {}
        finalTableDict['Attendance']['M-F'] = _monFri = round(timePerEvetTypeDict['attendance']['23456'] + layoverJoinUpDict['Layover']['23456'],2) #Needs to be summed to a discrepancy that is unknown
        finalTableDict['Attendance']['Sat'] = _sat = round(timePerEvetTypeDict['attendance']['7'] + layoverJoinUpDict['Layover']['7'],2) #Needs to be summed to a discrepancy that is unknown
        finalTableDict['Attendance']['Sun'] = _sun = round(timePerEvetTypeDict['attendance']['1'] + layoverJoinUpDict['Layover']['1'],2) #Needs to be summed to a discrepancy that is unknown
        finalTableDict['Attendance']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Attendance']['Comments'] = 'Recovery time whilst a driver is onboard'

        #Scheduled Bus Hours
        finalTableDict['Scheduled Bus Hours'] = {}
        finalTableDict['Scheduled Bus Hours']['M-F'] = _monFri = round(finalTableDict['In service Time']['M-F'] + finalTableDict['Deadtime']['M-F'] + finalTableDict['Attendance']['M-F'],2)
        finalTableDict['Scheduled Bus Hours']['Sat'] = _sat = round(finalTableDict['In service Time']['Sat'] + finalTableDict['Deadtime']['Sat'] + finalTableDict['Attendance']['Sat'],2)
        finalTableDict['Scheduled Bus Hours']['Sun'] = _sun = round(finalTableDict['In service Time']['Sun'] + finalTableDict['Deadtime']['Sun'] + finalTableDict['Attendance']['Sun'],2)
        finalTableDict['Scheduled Bus Hours']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Scheduled Bus Hours']['Comments'] = 'Total time drivers are in control of a vehicle'

        #Sign on
        finalTableDict['Sign on'] = {}
        finalTableDict['Sign on']['M-F'] = _monFri = round(timePerEvetTypeDict['sign_on']['23456'],2)
        finalTableDict['Sign on']['Sat'] = _sat = round(timePerEvetTypeDict['sign_on']['7'],2)
        finalTableDict['Sign on']['Sun']= _sun = round(timePerEvetTypeDict['sign_on']['1'],2)
        finalTableDict['Sign on']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Sign on']['Comments'] = 'Paid signing on time at start of duties'

        #Vehicle prep
        finalTableDict['Vehicle prep'] = {}
        finalTableDict['Vehicle prep']['M-F'] = _monFri = round(timePerEvetTypeDict['pre_trip']['23456'],2)
        finalTableDict['Vehicle prep']['Sat'] = _sat = round(timePerEvetTypeDict['pre_trip']['7'],2)
        finalTableDict['Vehicle prep']['Sun']= _sun = round(timePerEvetTypeDict['pre_trip']['1'],2)
        finalTableDict['Vehicle prep']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Vehicle prep']['Comments'] = 'Vehicle first user check before leaving depot'

        #Travel time
        finalTableDict['Travel time'] = {}
        finalTableDict['Travel time']['M-F'] = _monFri = round(timePerEvetTypeDict['public_travel']['23456'] + timePerEvetTypeDict['walk']['23456'] + timePerEvetTypeDict['relief_car']['23456'] + reliefTimeDict['23456'],2)
        finalTableDict['Travel time']['Sat'] = _sat = round(timePerEvetTypeDict['public_travel']['7'] + timePerEvetTypeDict['walk']['7'] + timePerEvetTypeDict['relief_car']['7'] + reliefTimeDict['7'],2)
        finalTableDict['Travel time']['Sun']= _sun = round(timePerEvetTypeDict['public_travel']['1'] + timePerEvetTypeDict['walk']['1'] + timePerEvetTypeDict['relief_car']['1'] + reliefTimeDict['1'],2)
        finalTableDict['Travel time']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Travel time']['Comments'] = 'Paid travel time for drivers'

        #Join Up
        finalTableDict['Join Up'] = {}
        finalTableDict['Join Up']['M-F'] = _monFri = round(layoverJoinUpDict['JOINUP']['23456'],2) #Needs to be summed to a discrepancy that is unknown
        finalTableDict['Join Up']['Sat'] = _sat = round(layoverJoinUpDict['JOINUP']['7'],2) #Needs to be summed to a discrepancy that is unknown
        finalTableDict['Join Up']['Sun'] = _sun = round(layoverJoinUpDict['JOINUP']['1'],2) #Needs to be summed to a discrepancy that is unknown
        finalTableDict['Join Up']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Join Up']['Comments'] = 'Paid time to join up work pieces'

        #Paid Breaks
        finalTableDict['Paid Breaks'] = {}
        finalTableDict['Paid Breaks']['M-F'] = _monFri = round(paidTimeDict['23456'],2)
        finalTableDict['Paid Breaks']['Sat'] = _sat = round(paidTimeDict['7'],2)
        finalTableDict['Paid Breaks']['Sun'] = _sun = round(paidTimeDict['1'],2)
        finalTableDict['Paid Breaks']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Paid Breaks']['Comments'] = 'Paid element of meal breaks over 45 mins and 2nd breaks (fully paid)'

        #CHANGED depot_pull_out to post_trip 
        #Vehicle Park
        finalTableDict['Vehicle Park'] = {}
        finalTableDict['Vehicle Park']['M-F'] = _monFri = round(timePerEvetTypeDict['post_trip']['23456'],2)
        finalTableDict['Vehicle Park']['Sat'] = _sat = round(timePerEvetTypeDict['post_trip']['7'],2)
        finalTableDict['Vehicle Park']['Sun']= _sun = round(timePerEvetTypeDict['post_trip']['1'],2)
        finalTableDict['Vehicle Park']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Vehicle Park']['Comments'] = 'Parking of bus when returning to depot'

       
        #Changed pre_trip to Sign_off
        #Sign off
        finalTableDict['Sign off'] = {}
        finalTableDict['Sign off']['M-F'] = _monFri = round(timePerEvetTypeDict['sign_off']['23456'],2)
        finalTableDict['Sign off']['Sat'] = _sat = round(timePerEvetTypeDict['sign_off']['7'],2)
        finalTableDict['Sign off']['Sun']= _sun = round(timePerEvetTypeDict['sign_off']['1'],2)
        finalTableDict['Sign off']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)

        finalTableDict['Make Up'] = {}
        finalTableDict['Make Up']['M-F'] = _monFri = round(makeupTimeDict['23456'],2)
        finalTableDict['Make Up']['Sat'] = _sat = round(makeupTimeDict['7'],2)
        finalTableDict['Make Up']['Sun']= _sun = round(makeupTimeDict['1'],2)
        finalTableDict['Make Up']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        
        
        
        

        
        
        #finalTableDict['Sign off']['Comments'] = 'Paying in time at end of shift'

        #Scheduled Paid Hours
        finalTableDict['Scheduled Paid Hours'] = {}
        _monFri = _sat = _sun = 0

        _relevantListSchedulePaidHours = list(set(list(finalTableDict.keys())) - set(['Scheduled Paid Hours', 'Scheduled Bus Hours', 'Pref Group']))

        for key in _relevantListSchedulePaidHours:
            _monFri += finalTableDict[key]['M-F']
            _sat += finalTableDict[key]['Sat']
            _sun += finalTableDict[key]['Sun']
        

        finalTableDict['Scheduled Paid Hours']['M-F'] = round(_monFri,2)
        finalTableDict['Scheduled Paid Hours']['Sat'] = round(_sat,2)
        finalTableDict['Scheduled Paid Hours']['Sun'] = round(_sun,2)
        finalTableDict['Scheduled Paid Hours']['Standard week'] = round(_monFri * 5 + _sat + _sun,2)
        #finalTableDict['Scheduled Paid Hours']['Comments'] = 'Total Scheduled paid time'

        #Scheduled Efficiency
        finalTableDict['Scheduled Efficiency'] = {}
        try:
            finalTableDict['Scheduled Efficiency']['M-F'] = round(finalTableDict['Scheduled Bus Hours']['M-F'] / finalTableDict['Scheduled Paid Hours']['M-F'], 3)*100
        except:
            finalTableDict['Scheduled Efficiency']['M-F'] = 0

        try:
            finalTableDict['Scheduled Efficiency']['Sat'] = round(finalTableDict['Scheduled Bus Hours']['Sat'] / finalTableDict['Scheduled Paid Hours']['Sat'],3)*100
        except:
            finalTableDict['Scheduled Efficiency']['Sat'] = 0

        try:
            finalTableDict['Scheduled Efficiency']['Sun'] = round(finalTableDict['Scheduled Bus Hours']['Sun'] / finalTableDict['Scheduled Paid Hours']['Sun'],3)*100
        except:
            finalTableDict['Scheduled Efficiency']['Sun'] = 0

        finalTableDict['Scheduled Efficiency']['Standard week'] = f"{round(finalTableDict['Scheduled Bus Hours']['Standard week'] / finalTableDict['Scheduled Paid Hours']['Standard week'],3)*100}%"
        #finalTableDict['Scheduled Efficiency']['Comments'] = '% of Paid time drivers are in control of a vehicle'

        
        #Percentage of Pay
       
        _relevantListPercentagePay = list(set(list(finalTableDict.keys())) - set(['Pref Group']))
        
        for key in _relevantListPercentagePay:
            if key in ['Scheduled Paid Hours', 'Scheduled Efficiency']:
                finalTableDict[key]['% of Pay'] = ''
            else:
                finalTableDict[key]['% of Pay'] = round(finalTableDict[key]['Standard week'] / finalTableDict['Scheduled Paid Hours']['Standard week'],3) *100

        #ASSIGNED COMMENTS AFTRER % OF PAY. Not the best solution but works...
        
        finalTableDict['In service Time']['Comments'] = 'In service driving time'
        finalTableDict['Deadtime']['Comments'] = 'Positioning trips time'
        finalTableDict['Attendance']['Comments'] = 'Recovery time whilst a driver is onboard'
        finalTableDict['Scheduled Bus Hours']['Comments'] = 'Total time drivers are in control of a vehicle'
        finalTableDict['Sign on']['Comments'] = 'Paid signing on time at start of duties'
        finalTableDict['Vehicle prep']['Comments'] = 'Vehicle first user check before leaving depot'
        finalTableDict['Travel time']['Comments'] = 'Paid travel time for drivers'
        finalTableDict['Join Up']['Comments'] = 'Paid time to join up work pieces'
        finalTableDict['Paid Breaks']['Comments'] = 'Paid element of meal breaks over 45 mins and 2nd breaks (fully paid)'
        finalTableDict['Vehicle Park']['Comments'] = 'Parking of bus when returning to depot'
        finalTableDict['Sign off']['Comments'] = 'Paying in time at end of shift'
        finalTableDict['Make Up']['Comments'] = 'Makeup Time'
        finalTableDict['Scheduled Paid Hours']['Comments'] = 'Total Scheduled paid time'
        finalTableDict['Scheduled Efficiency']['Comments'] = '% of Paid time drivers are in control of a vehicle'
        
        finalTableDict['Duty Fields'] = {}

        # st.write(dutyTypeDict)
        for dutyType in dutyTypeDict.keys():

            finalTableDict['Duty Fields'][dutyType] = {}
            for serviceId in dutyTypeDict[dutyType].keys():
                finalTableDict['Duty Fields'][dutyType][serviceId] = dutyTypeDict[dutyType][serviceId] #Getting the counters
            _monFri = finalTableDict['Duty Fields'][dutyType]['M-F']
            _sat = finalTableDict['Duty Fields'][dutyType]['Sat']
            _sun = finalTableDict['Duty Fields'][dutyType]['Sun']
            finalTableDict['Duty Fields'][dutyType]['Standard week'] = (_monFri * 5) + _sat + _sun
            finalTableDict['Duty Fields'][dutyType]['% of Pay'] = ''
            finalTableDict['Duty Fields'][dutyType]['Comments'] = 'Number of duties of type {}'.format(dutyType)
        
        # Calculations for the Totals

        dutyTypeList = list(dutyTypeDict.keys())
        serviceIdList = set(list(finalTableDict['Duty Fields'][dutyTypeList[0]].keys())) - set(['Comments', '% of Pay'])
        #serviceIdList = finalTableDict['Duty Fields'][dutyTypeList[0]].keys()

        finalTableDict['Total Duties'] = {}
        for serviceId in serviceIdList:
            finalTableDict['Total Duties'][serviceId] = 0
            for dutyType in dutyTypeList:
                finalTableDict['Total Duties'][serviceId] += finalTableDict['Duty Fields'][dutyType][serviceId]
        finalTableDict['Total Duties']['% of Pay'] = ''
        finalTableDict['Total Duties']['Comments'] = 'Total shift to cover each week'
    
        #Calculating the Mix fields
        finalTableDict['Duty Mix'] = {}
        for dutyType in dutyTypeDict.keys():
            newMixKeyName = f"{dutyType} Mix"
            finalTableDict['Duty Mix'][newMixKeyName] = {}
            for serviceId in serviceIdList:
                # "{0:.0%}".format
                try:
                    percentValue = finalTableDict['Duty Fields'][dutyType][serviceId] / finalTableDict['Total Duties'][serviceId]
                except:
                    percentValue = 0
                finalTableDict['Duty Mix'][newMixKeyName][serviceId] = "{0:.0%}".format(percentValue)
            finalTableDict['Duty Mix'][newMixKeyName]['% of Pay'] = ''
            finalTableDict['Duty Mix'][newMixKeyName]['Comments'] = '% Mix of {} duties'.format(dutyType)


        # st.write(finalTableDict)

        

        self.finalTableDict = finalTableDict

#get Pivot Table for Time per Event Type
def get_time_per_event_type_dict(data_frame, eventTypeList, serviceGroupDaysList):
    df = data_frame
    time_per_event_type_dict = {}

    for event_type in eventTypeList:
        time_per_event_type_dict[event_type] = {}
        for serviceGroup in serviceGroupDaysList:
            time_per_event_type_dict[event_type][serviceGroup] = df.loc[(df['Service Group Days'] == serviceGroup) & (df['Event Type'] == event_type), 'Time'].sum()

    timePerEvetTypeDict = time_per_event_type_dict

   
    return timePerEvetTypeDict

def get_makeup_per_duty_event(data_frame, serviceGroup):
    df = data_frame

    makeup_per_event_type_dict = {}

    for serviceGroup in serviceGroup:
        makeup_per_event_type_dict[serviceGroup] = df.loc[df['Service Group Days']== serviceGroup, 'makeup'].sum()

    return makeup_per_event_type_dict

#get Sheet 3 (aka Sum of Paid Time)
def get_pivot_sheet2(data_frame, serviceGroupDaysList):

    df = data_frame
    paidTimeDict = {}


    for serviceGroup in serviceGroupDaysList:
        paidTimeDict[serviceGroup] = df.loc[df['Service Group Days'] == serviceGroup, 'newPaid'].sum()

    return paidTimeDict

#get Sheet 5 (aka Sum of Layover / Join Up Value)
def get_pivot_sheet5(data_frame, serviceGroupDaysList):
    df = data_frame

    labelsList = list(set(df['Measure'].to_list()))
    layoverJoinUpDict = {}

    for measure in labelsList:
        layoverJoinUpDict[measure] = {}
        for serviceGroup in serviceGroupDaysList:
            layoverJoinUpDict[measure][serviceGroup] = df.loc[(df['Service Group Days'] == serviceGroup) & (df['Measure'] == measure), 'Layover / Join Up Value'].sum()
    
    return layoverJoinUpDict

#get Pivot Table Sum of Time per Relief Cars
def get_sum_of_time_relief_cars(data_frame, serviceGroupDaysList, _signList):

    _undesiredEventList = ['changeover', 'standby', 'split']

    df = data_frame
    df = df[~df['Event Type'].isin(_undesiredEventList)]
    df = df[df['Sign'].isin(_signList)]

    reliefTimeDict = {}

    for serviceGroup in serviceGroupDaysList:
        reliefTimeDict[serviceGroup] =  df.loc[df['Service Group Days'] == serviceGroup, 'Time'].sum()

    #Only first set in dictionary needs to be used
    
    return reliefTimeDict

#Get dictionary of number of duty Types
def get_duty_type_number_dict(data_frame):

    df = data_frame

    serviceIdList = list(set(df['Service Id']))

    dutyTypeDict = {}

    existingDutyTypesList = list(set(df['Type'].to_list()))

    for duty_type in existingDutyTypesList:
        dutyTypeDict[duty_type] = {}
        for service_id in serviceIdList:
            dutyTypeDict[duty_type][service_id] = df.loc[(df['Service Id'] == service_id) & (df['Type'] == duty_type), 'Type'].count()
    
    return dutyTypeDict

def check_and_update_dict(d):
    keys_to_check = [1, 23456, 7]
    if any(key not in d for key in keys_to_check):
        existing_key = next((key for key in d if key in keys_to_check), None)
        if existing_key is not None:
            subkeys = d[existing_key].keys()
            for key in keys_to_check:
                if key not in d:
                    d[key] = {subkey: 0 for subkey in subkeys}
        else:
            d[keys_to_check[0]] = {}
    return d

def check_subkeys(d):
    subkeys_list = ['1', '23456', '7']
    
    for key in d.keys():
        for subkey in subkeys_list:
            if subkey not in d[key]:
                d[key][subkey] = 0
    
    return d

def check_keys(d):
    keys_list = ['1', '23456', '7']
    
    for key in keys_list:
        if key not in d.keys():
            d[key] = 0

    return d

def add_missing_event_types(input_dict):

    eventTypeList = ['sign_on', 'pre_trip', 'depot_pull_out', 'service_trip',
                     'standby', 'changeover', 'attendance', 'walk', 'public_travel',
                     'sign_off', 'deadhead', 'depot_pull_in', 'post_trip', 'relief_car', 'split']
    
    for event in eventTypeList:
        if event not in input_dict:
            input_dict[event] = {}

    
    return input_dict


def check_subkeys_weekdays(d):
    subkeys_list = ['M-F', 'Sat', 'Sun']
    
    for key in d.keys():
        for subkey in subkeys_list:
            if subkey not in d[key]:
                d[key][subkey] = 0
    
    
    return d


#
#reliefTimeDict = get_sum_of_time_relief_cars(df, self.serviceGroupDaysList, _signList)

class TableBuilder():
    def __init__(self, file_path = None, AggrPath = None):
        self.file_path = file_path
        self.aggr_path = AggrPath
        self.fs = FullSchedule(self.file_path, self.aggr_path)
        self.fs.insert_extra_columns()

        

    def buildTableFile(self):
        prefGroupList = list(set(self.fs.adaptadedDataFrame['Pref Group'].to_list()))

        #print(prefGroupList)

        newPrefGroupList = [None]

        for item in prefGroupList:
            newPrefGroupList.append(item)

        #print(newPrefGroupList)
        
        tablesList = []

        timestamp = time.strftime("%Y%m%d%H%M%S")
        
        
        _fileName = "{}_Tables.xlsx".format(timestamp)

        

        #outputPath = os.path.join(_dirName,_fileName)
        output = BytesIO()
        wb = Workbook()
        #wb.save(output)

        
        #writer = pandas.ExcelWriter(output,engine='xlsxwriter')
        #df.to_excel(writer)
        #writer.save()
        #output.seek(0)
        #workbook = output.read()

        #wb = load_workbook(output)

        ws = wb.active

        for prefGroup in newPrefGroupList:
            self.fs.build_printable_table(prefGroup)
            newTable = self.fs.finalTableDict
            
            tablesList.append(newTable)
        
        self.tablesList = tablesList


        #print(tablesList)

        _headerList = ['Pay Category', 'M-F', 'Sat', 'Sun', 'Standard week', '% of Pay', 'Comments']
        _subKeyList = ['M-F', 'Sat', 'Sun', 'Standard week', '% of Pay', 'Comments']

        for index, _table in enumerate(self.tablesList):
            current_dict = self.tablesList[index]
            #st.write(current_dict)

            ws.append(['Summary of Paid Driver hours in a normal week'])
            ws.append(current_dict['Pref Group'])
            ws.append([''])
            ws.append(_headerList)

            for key in current_dict.keys():
                if key != 'Pref Group':
                    if key in ['Scheduled Efficiency', 'Duty Fields']:
                        ws.append([''])
                    if key == 'Duty Fields':
                        for dutyType in current_dict['Duty Fields'].keys():
                            _auxList = []
                            _auxList.append(dutyType)
                            for serviceId in current_dict['Duty Fields'][dutyType].keys():
                                _auxList.append(current_dict['Duty Fields'][dutyType][serviceId])
                            ws.append(_auxList)
                    elif key == 'Duty Mix':
                        dutyTypeList = list(current_dict['Duty Mix'].keys())
                        for dutyType in dutyTypeList:
                            _auxList = []
                            _auxList.append(dutyType)
                            # serviceIdList = list(current_dict['Duty Mix'][dutyType].keys())
                            serviceIdList = ['M-F', 'Sat', 'Sun', 'Standard week', '% of Pay', 'Comments']
                            for serviceId in serviceIdList:
                                _auxList.append(current_dict['Duty Mix'][dutyType][serviceId])
                            ws.append(_auxList)
                    else:
                        _auxList = []
                        _auxList.append(key)
                        for sub_item in _subKeyList:
                            _auxList.append(current_dict[key][sub_item])
                        ws.append(_auxList)

            ws.append([''])
            ws.append([''])
        
        # Styling the spreadsheet
        numOfTables = len(self.tablesList) # Storing the number of tables for iterating purposes
        rowsPerTable = 20 + 2* len(self.tablesList[0]['Duty Fields'].keys())

        totalRows = numOfTables * rowsPerTable

        #Alignments
        contentAlignment = Alignment(horizontal='center', vertical='center', wrap_text = False)
        rowTitleAlignment = Alignment(horizontal='left', vertical='center', wrap_text = False)

        #Fonts
        contentFont = Font(size = 11, bold = False)
        titleFont = Font(size = 14, bold = True, color='FFFFFF')

        #Fillings
        blueFilling = PatternFill(fill_type = 'solid', start_color = '366092', end_color = '366092')

        maxCol = 7
        minCol = 1

        blueTitleRows = ['Summary of Paid Driver hours in a normal week', 'Pay Category', 'Pay Category', 'Scheduled Bus Hours', 'Scheduled Paid Hours', 'Scheduled Efficiency']

        for row in ws.iter_rows(min_row=1, max_row=totalRows, min_col=minCol, max_col=maxCol):

            cell = ws.cell(row = row[0].row, column=1)

            if cell.value in blueTitleRows:

                for cell in row:
                    cell.font = titleFont
                    cell.alignment = contentAlignment
                    cell.fill = blueFilling
                    if cell.column in [1,7]:
                        cell.alignment = rowTitleAlignment


                if cell.value == "Summary of Paid Driver hours in a normal week":

                    ws.merge_cells(start_row=row[0].row, start_column=minCol, end_row=row[0].row, end_column=maxCol)

            else:

                for cell in row:
                    cell.font = contentFont
                    cell.alignment = contentAlignment
                    if cell.column in [1,7]:
                        cell.alignment = rowTitleAlignment
        
        wb.save(output)
        ste.download_button(label= 'Download Report', data=output, file_name=f'{file_name}'+f'{_fileName}')
st.subheader('First Bus Paid Time Report')
file_name = st.text_input('Output File Name', placeholder='Please enter the desired output file name (i.e: "Bolton")')
uploadedfile = st.file_uploader('Please select the Full Schedule File', type= 'xlsx')
uploadedAggrCrewSchedule = st.file_uploader('Please select the Aggregated Crew Schedule File', type= 'xlsx')

if uploadedfile and uploadedAggrCrewSchedule:    
    builder = TableBuilder(uploadedfile, uploadedAggrCrewSchedule)

    builder.buildTableFile()
    #builder.tablesList
    #fs = FullSchedule(uploadedfile)
    #fs.insert_extra_columns()
    #prefGroupList = ['Corridor']
    #fs.build_printable_table(prefGroupList)

    

    #dicio = fs.finalTableDict

    #for item in dicio.items():
        #print(item)




# ON the full schedule: 

#Total paid is time+joinup+paid+newpaid

# Move time to a new column called new_time, if it is a event_type =  standby or split don't use time but use paid value

# use matching key break and sum unique values to get total paid time 

#Check for values less 6 hr, makeup would be 6-value, 
